"""
make_contribution.py
--------------------
Generates Group63_Contribution.xlsx with the required columns:
    "Student Registration Number", "Name", "Percentage of contribution out of 100%"

Edit the STUDENTS list below to fill in each member's registration number and name.
Every member is set to 100% contribution (everyone contributed fully).

Run:
    python make_contribution.py
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

GROUP_ID = "63"
OUTPUT_FILE = f"Group{GROUP_ID}_Contribution.xlsx"

# --------------------------------------------------------------------------- #
# >>> EDIT THIS LIST: one tuple per student (Registration Number, Name) <<<
# All percentages are 100 because everyone contributed fully.
# --------------------------------------------------------------------------- #
STUDENTS = [
    ("<REG_NO_1>", "<Student Name 1>"),
    ("<REG_NO_2>", "<Student Name 2>"),
    ("<REG_NO_3>", "<Student Name 3>"),
    ("<REG_NO_4>", "<Student Name 4>"),
]

COLUMNS = [
    "Student Registration Number",
    "Name",
    "Percentage of contribution out of 100%",
]


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Group {GROUP_ID} Contribution"

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for r, (reg, name) in enumerate(STUDENTS, start=2):
        ws.cell(row=r, column=1, value=reg)
        ws.cell(row=r, column=2, value=name)
        c = ws.cell(row=r, column=3, value=100)
        c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 34
    return wb


if __name__ == "__main__":
    build_workbook().save(OUTPUT_FILE)
    print(f"Wrote {OUTPUT_FILE} with {len(STUDENTS)} students (all 100%).")
